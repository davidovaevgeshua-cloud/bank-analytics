"""
Обновление таблиц/графиков "Кредитный портфель юрлиц/физлиц" во вкладке
"Капитал" (form_capital/index.html) по официальной статистике ЦБ РФ
(показатель 4-1-3 "Отдельные показатели деятельности кредитных организаций").

Логика (аналогична scripts/update_data.py для CIR):
1. По data/loans_flul_state.json определяем следующий непроверенный период.
2. Пробуем скачать xlsx ЦБ за этот период (и ещё пару следующих, если ЦБ
   уже опубликовал несколько месяцев вперёд).
3. Если файл есть — парсим строку "Итого" (кредиты юрлицам / физлицам),
   дописываем в data/loans_flul_history.json.
4. Пересчитываем динамику (м/м, с начала года, г/г) по всей истории.
5. Подставляем свежий блок CBR_STAT в form_capital/index.html.

Период, отсутствующий на сайте ЦБ (404), — нормальная ситуация: значит,
статистика за этот месяц ещё не вышла. Скрипт останавливается на первом
таком периоде и завершается без ошибки.
"""
import json
import re
from datetime import date
from pathlib import Path
from urllib.request import Request, urlopen
from urllib.error import HTTPError

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
HISTORY_PATH = ROOT / "data" / "loans_flul_history.json"
STATE_PATH = ROOT / "data" / "loans_flul_state.json"
TARGET_HTML = ROOT / "form_capital" / "index.html"
TMP_DIR = ROOT / "_tmp_loans_flul_download"

BASE_URL = "https://www.cbr.ru/vfs/statistics/pdko/pdko_sub/perf-ind/stat_bn_4-1-3_01{mm:02d}{yyyy:04d}.xlsx"
USER_AGENT = "Mozilla/5.0 (compatible; bank-analytics-bot/1.0)"

MAX_PERIODS_PER_RUN = 3
BASE_PERIOD = "122023"  # якорь ряда (дек. 2023), данных за этот период нет — всегда null


def load_json(path: Path, default):
    if path.exists():
        return json.loads(path.read_text(encoding="utf-8"))
    return default


def save_json(path: Path, data):
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(data, ensure_ascii=False, indent=1), encoding="utf-8")


def next_period(label: str) -> str:
    mm, yyyy = int(label[:2]), int(label[2:])
    mm += 1
    if mm > 12:
        mm = 1
        yyyy += 1
    return f"{mm:02d}{yyyy}"


def period_to_iso(label: str) -> str:
    """'012024' -> '2024-01'."""
    return f"{label[2:]}-{label[:2]}"


def download_xlsx(period_label: str, dest_dir: Path):
    mm, yyyy = int(period_label[:2]), int(period_label[2:])
    url = BASE_URL.format(mm=mm, yyyy=yyyy)
    dest_dir.mkdir(parents=True, exist_ok=True)
    out_path = dest_dir / f"stat_bn_4-1-3_{period_label}.xlsx"
    req = Request(url, headers={"User-Agent": USER_AGENT})
    try:
        with urlopen(req, timeout=60) as resp, open(out_path, "wb") as f:
            f.write(resp.read())
    except HTTPError as e:
        if e.code == 404:
            return None
        raise
    return out_path


def parse_xlsx(path: Path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    loans_corp = loans_retail = None
    for row in ws.iter_rows(min_row=1, max_row=30, values_only=True):
        label = row[0]
        if not label:
            continue
        total = row[7] if len(row) > 7 else None
        if label.startswith("Корпоративные кредиты"):
            loans_corp = total
        elif label.startswith("Кредиты физическим лицам"):
            loans_retail = total
    if loans_corp is None or loans_retail is None:
        raise RuntimeError(f"Не нашёл ожидаемые строки в {path.name} (структура файла ЦБ изменилась?)")
    return loans_corp, loans_retail


def build_periods_full(history: dict) -> list:
    labels = sorted(history.keys(), key=lambda p: (int(p[2:]), int(p[:2])))
    last = labels[-1]
    periods = [BASE_PERIOD]
    label = next_period(BASE_PERIOD)
    while True:
        periods.append(label)
        if label == last:
            break
        label = next_period(label)
    return periods


def growth_series(vals, periods_full):
    mom, ytd, yoy = [], [], []
    idx = {p: i for i, p in enumerate(periods_full)}
    for i, p in enumerate(periods_full):
        v = vals[i]
        prev = vals[i - 1] if i > 0 else None
        mom.append(round((v / prev - 1) * 100, 2) if v is not None and prev not in (None, 0) else None)

        mm, yyyy = int(p[:2]), int(p[2:])
        dec_prev = f"12{yyyy - 1}"
        j = idx.get(dec_prev)
        base = vals[j] if j is not None else None
        ytd.append(round((v / base - 1) * 100, 2) if v is not None and base not in (None, 0) else None)

        base12 = vals[i - 12] if i >= 12 else None
        yoy.append(round((v / base12 - 1) * 100, 2) if v is not None and base12 not in (None, 0) else None)
    return mom, ytd, yoy


def build_cbr_stat(history: dict) -> dict:
    periods_full = build_periods_full(history)
    loans_corp = []
    loans_retail = []
    for p in periods_full:
        e = history.get(p)
        if e is None:
            loans_corp.append(None)
            loans_retail.append(None)
        else:
            loans_corp.append(round(e["loans_corp"] / 1e6, 3))
            loans_retail.append(round(e["loans_retail"] / 1e6, 3))

    corp_mom, corp_ytd, corp_yoy = growth_series(loans_corp, periods_full)
    retail_mom, retail_ytd, retail_yoy = growth_series(loans_retail, periods_full)

    return {
        "periods": [period_to_iso(p) for p in periods_full],
        "loans_corp": loans_corp,
        "loans_retail": loans_retail,
        "corp_mom": corp_mom, "corp_ytd": corp_ytd, "corp_yoy": corp_yoy,
        "retail_mom": retail_mom, "retail_ytd": retail_ytd, "retail_yoy": retail_yoy,
        "source_note": "Форма стат. ЦБ РФ 4-1-3 «Отдельные показатели деятельности кредитных организаций», раздел «Итого» по сектору. Кредиты юрлицам — с учётом ИП и приобретённых прав требования; кредиты физлицам — аналогично.",
    }


def render(cbr_stat: dict):
    if not TARGET_HTML.exists():
        raise RuntimeError(f"Не найден {TARGET_HTML} — вкладка «Капитал» ещё не установлена в репозиторий?")
    html = TARGET_HTML.read_text(encoding="utf-8")
    js_json = json.dumps(cbr_stat, ensure_ascii=False)
    new_html, n = re.subn(r"const CBR_STAT = \{.*?\};", f"const CBR_STAT = {js_json};", html, flags=re.DOTALL)
    if n != 1:
        raise RuntimeError("Не нашёл блок 'const CBR_STAT = {...};' в form_capital/index.html — структура файла изменилась?")
    TARGET_HTML.write_text(new_html, encoding="utf-8")


def main():
    state = load_json(STATE_PATH, {"last_period": "062026"})
    history = load_json(HISTORY_PATH, {})
    processed = []

    label = next_period(state["last_period"])
    for _ in range(MAX_PERIODS_PER_RUN):
        today = date.today()
        yyyy, mm = int(label[2:]), int(label[:2])
        if (yyyy, mm) > (today.year, today.month):
            break
        path = download_xlsx(label, TMP_DIR)
        if path is None:
            print(f"Период {label} пока не опубликован ЦБ — останавливаюсь здесь.")
            break
        loans_corp, loans_retail = parse_xlsx(path)
        history[label] = {"loans_corp": loans_corp, "loans_retail": loans_retail}
        processed.append(label)
        state["last_period"] = label
        label = next_period(label)

    if processed:
        save_json(HISTORY_PATH, history)
        save_json(STATE_PATH, state)
        cbr_stat = build_cbr_stat(history)
        render(cbr_stat)
        print(f"Обновлены периоды: {processed}. {TARGET_HTML} пересобран.")
    else:
        print("Новых периодов не найдено, обновление не требуется.")

    if TMP_DIR.exists():
        import shutil
        shutil.rmtree(TMP_DIR)


if __name__ == "__main__":
    main()
